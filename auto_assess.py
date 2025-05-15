# Copyright (c) Microsoft. All rights reserved.
# Licensed under the MIT license. See LICENSE.md file in the project root for full license information.


import openpyxl
from openai import AzureOpenAI
from azure.search.documents import SearchClient

import os
import sys
import argparse
from dotenv import load_dotenv
from azure.identity import DefaultAzureCredential
import json

load_dotenv(override=True)

OPENAI_ENDPOINT = os.getenv("AZURE_OPENAI_ENDPOINT")
#OPENAI_DEPLOYMENT = os.getenv("AZURE_OPENAI_DEPLOYMENT")
OPENAI_DEPLOYMENT = "global-4o"
OPENAI_REASONING_DEPLOYMENT = os.getenv("AZURE_OPENAI_REASONING_DEPLOYMENT")
SEARCH_ENDPOINT = os.getenv("AZURE_AI_SEARCH_ENDPOINT")
SEARCH_INDEX = os.getenv("AZURE_AI_SEARCH_INDEX")
AZURE_OPENAI_EMBEDDING_DEPLOYMENT = os.getenv("AZURE_OPENAI_EMBEDDING_DEPLOYMENT")

print(f"OPENAI_ENDPOINT: {OPENAI_ENDPOINT}")
print(f"OPENAI_DEPLOYMENT: {OPENAI_DEPLOYMENT}")                
print(f"SEARCH_ENDPOINT: {SEARCH_ENDPOINT}")
print(f"SEARCH_INDEX: {SEARCH_INDEX}")
print(f"AZURE_OPENAI_EMBEDDING_DEPLOYMENT: {AZURE_OPENAI_EMBEDDING_DEPLOYMENT}")

# Initialize global variables for column overrides and prompt rules
COLUMN_OVERRIDES = {}
PROMPT_RULES = {}
SYSTEM_PROMPT = ""


# Load configurations from JSON
def load_config(json_filepath):
    with open(json_filepath, 'r', encoding='utf-8') as f:
        return json.load(f)

# Read text from a file
def read_text_from_file(file_path):
    """Read text content from a file."""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read().strip()
    except Exception as e:
        print(f"Error reading file {file_path}: {e}")
        return None


# Initialize configuration from project directory
def init_config(project_dir):
    global COLUMN_OVERRIDES, PROMPT_RULES, SYSTEM_PROMPT
    # Get paths to configuration files
    column_override_path = os.path.join(project_dir, "column_overrides.json")
    prompt_rules_path = os.path.join(project_dir, "prompt_rules.json")
    
    # Get the base name of the project directory without parent folders
    project_base_name = os.path.basename(project_dir.rstrip(os.path.sep))
    # Remove "_project" suffix if present
    project_prompts_base = project_base_name.replace("_project", "")
    project_prompts_dir = os.path.join(project_dir, f"{project_prompts_base}_prompts")
    system_prompt_path = os.path.join(project_prompts_dir, "system_prompt.json")
    
    print("Project directory:", project_dir)
    print("Project base name:", project_base_name)
    print("Column overrides path:", column_override_path)
    print("Prompt rules path:", prompt_rules_path)
    print("System prompt path:", system_prompt_path)

    # Load column overrides
    if os.path.exists(column_override_path):
        COLUMN_OVERRIDES = load_config(column_override_path)
        print(f"Loaded COLUMN_OVERRIDES: {COLUMN_OVERRIDES}")
    else:
        print(f"Warning: Column overrides file not found at {column_override_path}")

    # Load prompt rules
    if os.path.exists(prompt_rules_path):
        PROMPT_RULES = load_config(prompt_rules_path)
        print(f"Loaded PROMPT_RULES: {PROMPT_RULES}")
    else:
        print(f"Warning: Prompt rules file not found at {prompt_rules_path}")

    # Load system prompt from JSON
    if os.path.exists(system_prompt_path):
        system_prompt_data = load_config(system_prompt_path)
        # Handle different possible JSON structures
        if isinstance(system_prompt_data, str):
            # If the file contains just a string
            SYSTEM_PROMPT = system_prompt_data
        elif isinstance(system_prompt_data, dict):
            # Try different common keys that might contain the system prompt
            if "system_prompt" in system_prompt_data:
                SYSTEM_PROMPT = system_prompt_data["system_prompt"]
            elif "prompt" in system_prompt_data:
                SYSTEM_PROMPT = system_prompt_data["prompt"]
            elif "content" in system_prompt_data:
                SYSTEM_PROMPT = system_prompt_data["content"]
            elif "text" in system_prompt_data:
                SYSTEM_PROMPT = system_prompt_data["text"]
            else:
                # If no known keys, convert the entire JSON to a string
                SYSTEM_PROMPT = json.dumps(system_prompt_data, indent=2)
        else:
            # For any other type, convert to string
            SYSTEM_PROMPT = str(system_prompt_data)
        
        print(f"Loaded SYSTEM_PROMPT from JSON: {SYSTEM_PROMPT[:100]}..." if len(SYSTEM_PROMPT) > 100 else f"Loaded SYSTEM_PROMPT from JSON: {SYSTEM_PROMPT}")
    else:
        print(f"Warning: System prompt file not found at {system_prompt_path}")

def append_grading_rubric(rubric_path):
    """Append grading rubric to the system prompt if provided"""
    global SYSTEM_PROMPT
    
    if rubric_path and os.path.exists(rubric_path):
        rubric_content = read_text_from_file(rubric_path)
        if rubric_content:
            SYSTEM_PROMPT += f"\n\nGrading Rubric:\n{rubric_content}"
            print(f"Appended grading rubric to system prompt: {SYSTEM_PROMPT}")
        else:
            print(f"Warning: Could not read grading rubric from {rubric_path}")
    elif rubric_path:
        print(f"Warning: Grading rubric file not found at {rubric_path}")

def col_header_override(col_header: str) -> str:
    normalized = ' '.join(col_header.lower().split())
    # Return override if it exists
    return COLUMN_OVERRIDES.get(normalized, col_header)

def get_specific_prompt(col_header: str, search_query: str, context: str) -> str:
    # Normalize the column header
    normalized = ' '.join(col_header.lower().split())
    # Look up the prompt rule
    print("get_specific_prompt: Normalized column header:", normalized)
    print("get_specific_prompt: ",PROMPT_RULES)
    print("get_specific_prompt: search_query:", search_query)

    prompt_template = PROMPT_RULES.get(normalized)
    print("get_specific_prompt: Prompt template:", prompt_template)
    if prompt_template:
        # Fill placeholders
        col_prompt = prompt_template.replace("{search_query}", search_query).replace("{context}", context)
    else:
        # Fallback default
        col_prompt = f"provide verbatim response for {search_query} based on the following information: {context}"

    return col_prompt




# Use DefaultAzureCredential for authentication
credential = DefaultAzureCredential()

# Create the Azure Search client
search_client = SearchClient(endpoint=SEARCH_ENDPOINT, index_name=SEARCH_INDEX, credential=credential)

# Initialize Azure OpenAI client with DefaultAzureCredential
client = AzureOpenAI(
    azure_endpoint=OPENAI_ENDPOINT,
    api_version="2023-05-15",
    azure_ad_token=credential.get_token("https://cognitiveservices.azure.com/.default").token
)

# Function to fetch relevant context from Azure Search using hybrid search with semantic ranking
def azure_search_mm_doc_proc_rag(query: str) -> str:
    """
    Function to fetch relevant context from Azure Search using hybrid search with semantic ranking.
    Returns the text field from the top-ranked search result in @search.answers.
    """
    print(f"\n azure_search_rag: Query: {query}\n", flush=True)
    # Generate embedding vector from Azure OpenAI
    embedding_response = client.embeddings.create(
        model=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        input=query
    )
    embedding_vector = embedding_response.data[0].embedding

    # Corrected vector query with kind="vector" as per Explorer payload
    search_body = {
        "search": query,
        "top": 3,  # Get top 3 results
        "count": True,
        "queryLanguage": "en-us",
        "queryType": "semantic",
        "semanticConfiguration": "my-semantic-config",
        "answers": "extractive|count-3",
        "captions": "extractive",
        "vectorQueries": [
            {
                "kind": "text",
                "text": query,
                "fields": "text_vector"
            }
        ],
        "searchFields": "text",
        "searchMode": "all"
    }

    # Perform the search using the lower-level client method
    response = search_client._client.documents.search_post(search_body)
    raw_response = response.as_dict()
   # print(f"azure_search_rag: Raw response: {json.dumps(raw_response, indent=2)}")
    #"text_vector"
    
    context = "No relevant information found."
    
    # Prefer the "results" section based on highest reranker_score
    if 'results' in raw_response and raw_response['results']:
        top_result = max(raw_response['results'], key=lambda r: r.get('reranker_score', 0))
        text = top_result.get('text')
        if text:
            context = text
            print(f"azure_search_rag: Top result based on reranker_score: {context}")
            return context

    # Fallback: check if there's an answer available
    if 'answers' in raw_response and raw_response['answers']:
        answers = raw_response['answers']
        if len(answers) > 0 and 'text' in answers[0]:
            context = answers[0]['text']
            print(f"azure_search_rag: Top semantic answer text: {context}")
            return context

    # Return if no result found
    return context


def azure_search_rag_this(query: str) -> str:
    """
    Function to fetch relevant context from Azure Search using hybrid search with semantic ranking.
    This function expects the search index to be configured based on a simple schema as creaded by the "rag_this" libraries 
    with a text_vector field for vector search, and the text field containing the searchable text (in markdown format).)
    Returns the text field from the top-ranked search result in @search.answers.
    """
    print(f"\n azure_search_rag: Query: {query}\n", flush=True)
    # Generate embedding vector from Azure OpenAI
    embedding_response = client.embeddings.create(
        model=AZURE_OPENAI_EMBEDDING_DEPLOYMENT,
        input=query
    )
    embedding_vector = embedding_response.data[0].embedding

    # Corrected vector query with kind="vector" as per Explorer payload
    search_body = {
        "search": query,
        "top": 3,  # Get top 3 results
        "count": True,
        "queryLanguage": "en-us",
        "queryType": "semantic",
        "semanticConfiguration": "vector-semantic-configuration",
        "answers": "extractive|count-3",
        "captions": "extractive",
        "vectorQueries": [
            {
                "kind": "text",
                "text": query,
                "fields": "text_vector"
            }
        ],
        "searchFields": "text",
        "searchMode": "all"
    }

    # Perform the search using the lower-level client method
    response = search_client._client.documents.search_post(search_body)
    raw_response = response.as_dict()
    # Remove the print of raw response which contains vector data
    # print(f"azure_search_rag: Raw response: {json.dumps(raw_response, indent=2)}")
    
    context = "No relevant information found."
    
    # Prefer the "results" section based on highest reranker_score
    if 'results' in raw_response and raw_response['results']:
        top_result = max(raw_response['results'], key=lambda r: r.get('reranker_score', 0))
        text = top_result.get('text')
        if text:
            context = text
            print(f"azure_search_rag: Found relevant content")
            return context

    # Fallback: check if there's an answer available
    if 'answers' in raw_response and raw_response['answers']:
        answers = raw_response['answers']
        if len(answers) > 0 and 'text' in answers[0]:
            context = answers[0]['text']
            print(f"azure_search_rag: Found answer from semantic search")
            return context

    # Return if no result found
    return context

def call_openai(prompt: str) -> str:
    """
    Function to call Azure OpenAI with your RAG context.
    """
    # Only print the first 200 characters of the prompt for debugging
    print("call_openai: Prompt:", prompt[:1000])
    response = client.chat.completions.create(
        model=OPENAI_DEPLOYMENT,
        messages=[
            {"role": "system", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt}
        ],
        max_tokens=1000,  
        temperature=0,  
        top_p=0.95,  
        frequency_penalty=0,  
        presence_penalty=0,  
        stop=None,  
    )
    return response.choices[0].message.content.strip()

def call_openai_reasoning(prompt: str) -> str:
    """
    Function to call Azure OpenAI with your RAG context.
    """
    # Only print the first 200 characters of the prompt for debugging
    print("call_openai: Prompt:", prompt[:1000])
    response = client.chat.completions.create(
        model=OPENAI_REASONING_DEPLOYMENT,
        messages=[
            {"role": "user", "content": SYSTEM_PROMPT},
            {"role": "user", "content": prompt}
        ],
        max_tokens=20000,  
        temperature=0,  
        reasoning_effort="high",  
        stop=None,  
    )
    return response.choices[0].message.content.strip()



# Modify process_cell to use a specific prompt for each column
def process_cell(system_prompt: str, row_header: str, col_header: str) -> str:
    col_header = col_header_override(col_header) # Override the column header for specific prompts
    
    # Convert row_header to string if it's not already a string
    row_header_str = str(row_header) if row_header is not None else ""
    
    # Check if SEARCH_QUERY_CONFIG environment variable is set
    SearchQuery = os.getenv("SEARCH_QUERY_CONFIG")
    if not SearchQuery:
        # Fallback to default if environment variable not set
        SearchQuery = col_header + ' question number ' + row_header_str
    else:
        # Replace placeholders with actual values
        SearchQuery = SearchQuery.replace("{col_header}", col_header).replace("{row_header}", row_header_str)
    
    print(f"process_cell: SearchQuery: {SearchQuery}\n", flush=True)

    #context = azure_search_mm_doc_proc_rag(SearchQuery)
    context = azure_search_rag_this(SearchQuery)
    specific_prompt = get_specific_prompt(col_header, SearchQuery, context)
    prompt = f"{system_prompt}\n{specific_prompt}"
    return call_openai(prompt)

def populate_cells_in_workbook(filepath: str, sheet_name: str, start_row: int, end_row: int, start_col: int, end_col: int):
    wb = openpyxl.load_workbook(filepath)
    sheet = wb[sheet_name]

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            row_header = sheet.cell(row=row, column=1).value
            col_header = sheet.cell(row=1, column=col).value

            # Process the cell directly since we're not using a formal flow
            print(f"Processing cell ({row}, {col}) with row header: {row_header}, col header: {col_header}\n",flush=True)
            cell_value = process_cell(SYSTEM_PROMPT, row_header, col_header)
            print(f"Cell value response: {cell_value}\n")
            sheet.cell(row=row, column=col).value = cell_value
    
    # Reverted to original saving method
    wb.save(filepath)

# Replace the previous check_file_writeable with a function that uses os.rename to check if the file is unlocked.
def check_file_writeable(filepath: str):
    import os, time
    while True:
        try:
            # Attempt to rename the file to itself; if the file is locked, this will raise a PermissionError.
            os.rename(filepath, filepath)
            return
        except PermissionError as e:
            print(f"Error: File {filepath} is currently in use and cannot be written to (Permission denied).")
            user_input = input("Please close the file and type OK to retry or type Cancel to exit: ").strip().lower()
            if user_input == "cancel":
                print("Exiting the program.")
                exit(1)
            else:
                print("Retrying in 2 seconds...")
                time.sleep(2)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Process APP spreadsheet using Azure OpenAI and Azure Search.")
    parser.add_argument("--target-file", required=True, help="Path to the target Excel file.")
    parser.add_argument("--project-dir", required=True, help="Path to the project directory containing configuration files.")
    parser.add_argument("--grading-rubric", help="Optional path to a grading rubric file.")
    parser.add_argument("--sheet-name", default="Sheet1", help="Name of the sheet to process (default: Sheet1)")
    parser.add_argument("--start-row", type=int, default=2, help="Starting row number (default: 2)")
    parser.add_argument("--end-row", type=int, default=2, help="Ending row number (default: 2)")
    parser.add_argument("--start-col", type=int, default=6, help="Starting column number (default: 6)")
    parser.add_argument("--end-col", type=int, default=6, help="Ending column number (default: 6)")
    args = parser.parse_args()

    target_filepath = args.target_file
    project_dir = args.project_dir
    grading_rubric_path = args.grading_rubric

    # Initialize configuration from project directory
    init_config(project_dir)
    
    # Append grading rubric to system prompt if provided
    if grading_rubric_path:
        append_grading_rubric(grading_rubric_path)

    # Print configuration summary
    print("\nConfiguration Summary:")
    print(f"Target Excel file: {target_filepath}")
    print(f"Project directory: {project_dir}")
    print(f"Sheet name: {args.sheet_name}")
    print(f"Cell range: ({args.start_row},{args.start_col}) to ({args.end_row},{args.end_col})")
    if grading_rubric_path:
        print(f"Grading rubric: {grading_rubric_path}")
    print("\n")

    # Check that the target file is writeable before processing; if not, ask user to close it.
    check_file_writeable(target_filepath)
    
    populate_cells_in_workbook(
        filepath=target_filepath,
        sheet_name=args.sheet_name,
        start_row=args.start_row,
        end_row=args.end_row,
        start_col=args.start_col,
        end_col=args.end_col
    )
