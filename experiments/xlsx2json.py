#!/usr/bin/env python3

# Copyright (c) Microsoft. All rights reserved.
# Licensed under the MIT license. See LICENSE.md file in the project root for full license information.


"""
Excel to JSON and JSON to Excel utilities for AI-based cell completion.

This module provides functions to:
1. Extract data from specific Excel ranges
2. Create JSON templates based on Excel headers and data
3. Update Excel spreadsheets with values from populated JSON
"""

import json
import os
import re
from typing import Dict, List, Any, Tuple, Optional, Union
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string

class ExcelJsonProcessor:
    def __init__(self, excel_path: str):
        """Initialize with path to Excel file."""
        self.excel_path = excel_path
        self.workbook = None
        self._load_workbook()
    
    def _load_workbook(self):
        """Load the Excel workbook."""
        if os.path.exists(self.excel_path):
            self.workbook = openpyxl.load_workbook(self.excel_path)
        else:
            raise FileNotFoundError(f"Excel file not found: {self.excel_path}")
    
    def parse_cell_range(self, range_str: str) -> Tuple[str, str, int, str, int]:
        """
        Parse a cell range string like 'Sheet1!A1:C10' or 'A1:C10'.
        
        Returns:
            Tuple of (sheet_name, start_col, start_row, end_col, end_row)
        """
        # Check if sheet name is included
        if '!' in range_str:
            sheet_name, cell_range = range_str.split('!')
        else:
            sheet_name = self.workbook.active.title
            cell_range = range_str
        
        # Parse the range
        start, end = cell_range.split(':')
        start_coord = coordinate_from_string(start)
        end_coord = coordinate_from_string(end)
        
        start_col, start_row = start_coord
        end_col, end_row = end_coord
        
        return sheet_name, start_col, start_row, end_col, end_row

    def get_headers_from_range(self, range_str: str, header_row: int = None) -> List[str]:
        """
        Extract headers from a specific row in the range.
        
        Args:
            range_str: Excel range like 'Sheet1!A1:C10'
            header_row: Row number containing headers (if None, uses first row of range)
        
        Returns:
            List of header names
        """
        sheet_name, start_col, start_row, end_col, end_row = self.parse_cell_range(range_str)
        sheet = self.workbook[sheet_name]
        
        # Determine header row
        if header_row is None:
            header_row = start_row
        
        # Get column indices
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        # Extract headers
        headers = []
        for col_idx in range(start_col_idx, end_col_idx + 1):
            col_letter = get_column_letter(col_idx)
            cell_value = sheet[f"{col_letter}{header_row}"].value
            # Clean up header for use as JSON key
            if cell_value:
                # Convert to string and clean for JSON key use
                header = str(cell_value).strip()
                # Replace spaces with underscores and remove special characters
                header = re.sub(r'[^\w\s]', '', header)
                header = header.replace(' ', '_').lower()
                headers.append(header)
            else:
                headers.append(f"column_{col_letter}")
        
        return headers

    def extract_data_from_range(self, range_str: str, include_headers: bool = True) -> Dict[str, List[Any]]:
        """
        Extract data from a specified range and organize by columns.
        
        Args:
            range_str: Excel range like 'Sheet1!A1:C10'
            include_headers: Whether to include headers row in the data
        
        Returns:
            Dictionary with column headers as keys and column values as lists
        """
        sheet_name, start_col, start_row, end_col, end_row = self.parse_cell_range(range_str)
        sheet = self.workbook[sheet_name]
        
        # Get headers first
        headers = self.get_headers_from_range(range_str)
        
        # Initialize data structure
        data = {header: [] for header in headers}
        
        # Determine start row for data (skip header if needed)
        data_start_row = start_row + 1 if not include_headers else start_row
        
        # Get column indices
        start_col_idx = column_index_from_string(start_col)
        end_col_idx = column_index_from_string(end_col)
        
        # Extract data by columns
        for row in range(data_start_row, end_row + 1):
            for col_idx, header in zip(range(start_col_idx, end_col_idx + 1), headers):
                col_letter = get_column_letter(col_idx)
                cell_value = sheet[f"{col_letter}{row}"].value
                data[header].append(cell_value)
        
        return data

    def create_json_template(self, range_str: str, description: str = "", 
                            exclude_columns: List[str] = None, 
                            multi_row: bool = True) -> Dict[str, Any]:
        """
        Create a JSON template with headers as keys and empty placeholders as values.
        
        Args:
            range_str: Excel range like 'Sheet1!A1:C10'
            description: Description of the template for AI context
            exclude_columns: List of column headers to exclude
            multi_row: Whether to create a template for multiple rows (True) or just one row (False)
        
        Returns:
            JSON template dictionary
        """
        sheet_name, start_col, start_row, end_col, end_row = self.parse_cell_range(range_str)
        headers = self.get_headers_from_range(range_str)
        
        if exclude_columns:
            headers = [h for h in headers if h not in exclude_columns]
        
        # Create a template with descriptive placeholder for each header
        template = {
            "description": description,
            "range": range_str,
            "data": {}
        }
        
        # If multi_row is True, create a list of placeholders for each header
        # Otherwise, create a single placeholder for each header
        num_rows = end_row - start_row
        if num_rows > 1 and multi_row:
            for header in headers:
                # Create a list of placeholders for each row
                template["data"][header] = [
                    f"[AI to provide {header.replace('_', ' ')} for row {i+1}]" 
                    for i in range(num_rows)
                ]
        else:
            # Create a single placeholder for each header (original behavior)
            for header in headers:
                template["data"][header] = f"[AI to provide {header.replace('_', ' ')}]"
        
        return template

    def save_json_template(self, template: Dict[str, Any], output_path: str) -> None:
        """Save the JSON template to a file."""
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(template, f, indent=2)
        print(f"JSON template saved to {output_path}")

    def load_populated_json(self, json_path: str) -> Dict[str, Any]:
        """Load a populated JSON file."""
        with open(json_path, 'r', encoding='utf-8') as f:
            return json.load(f)

    def update_excel_from_json(self, json_data: Dict[str, Any], target_range: str = None) -> None:
        """
        Update Excel cells with values from populated JSON.
        
        Args:
            json_data: The populated JSON data
            target_range: Optional override for the range in the JSON
        """
        # Get the range from JSON or use the provided override
        excel_range = target_range if target_range else json_data.get("range")
        if not excel_range:
            raise ValueError("No target range specified in JSON or as parameter")
        
        sheet_name, start_col, start_row, end_col, end_row = self.parse_cell_range(excel_range)
        sheet = self.workbook[sheet_name]
        
        # Get headers for the range to map JSON keys to columns
        headers = self.get_headers_from_range(excel_range)
        
        # Get the data from JSON
        json_values = json_data.get("data", {})
        
        # Check that our JSON keys match the headers
        for header in headers:
            if header not in json_values and header in json_data.get("data", {}):
                print(f"Warning: Header '{header}' from Excel not found in JSON")
        
        # Map JSON data to Excel columns
        start_col_idx = column_index_from_string(start_col)
        
        # Start from the row after the header (if we're treating the first row as headers)
        current_row = start_row + 1
        
        # For each JSON key, update the corresponding Excel column
        for col_idx, header in enumerate(headers, start=start_col_idx):
            if header in json_values:
                col_letter = get_column_letter(col_idx)
                value = json_values[header]
                
                # Handle different types of values
                if isinstance(value, list):
                    # If the value is a list, update multiple rows
                    for i, item in enumerate(value):
                        if current_row + i <= end_row:  # Stay within range
                            sheet[f"{col_letter}{current_row + i}"] = item
                else:
                    # Single value - use just for the first row after header
                    sheet[f"{col_letter}{current_row}"] = value
        
        # Save the updated workbook
        self.workbook.save(self.excel_path)
        print(f"Excel file updated at {self.excel_path}")


def main():
    """Example usage of the ExcelJsonProcessor."""
    import argparse
    
    parser = argparse.ArgumentParser(description="Excel to JSON and JSON to Excel utilities")
    parser.add_argument('--excel', required=True, help="Path to the Excel file")
    parser.add_argument('--range', required=True, help="Cell range (e.g., 'Sheet1!A1:C10')")
    parser.add_argument('--output', help="Output JSON template path (required for extract mode)")
    parser.add_argument('--description', default="", help="Description for the JSON template")
    parser.add_argument('--mode', choices=['extract', 'update'], default='extract', 
                        help="Mode: extract (Excel to JSON) or update (JSON to Excel)")
    parser.add_argument('--json', help="Path to populated JSON (for update mode)")
    parser.add_argument('--multi-row', action='store_true', default=True,
                        help="Create JSON template with multiple rows")
    
    args = parser.parse_args()
    
    processor = ExcelJsonProcessor(args.excel)
    
    if args.mode == 'extract':
        # Check if output path is provided for extract mode
        if not args.output:
            parser.error("--output argument is required for extract mode")
            
        # Create JSON template from Excel
        template = processor.create_json_template(args.range, args.description, multi_row=args.multi_row)
        processor.save_json_template(template, args.output)
    elif args.mode == 'update':
        if not args.json:
            parser.error("--json argument is required for update mode")
        
        # Load populated JSON and update Excel
        json_data = processor.load_populated_json(args.json)
        processor.update_excel_from_json(json_data)


if __name__ == "__main__":
    main()