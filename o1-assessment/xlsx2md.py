import openpyxl
import os
import argparse
from pathlib import Path


def xlsx_to_markdown(file_path, output_dir="markdown_output"):
    """
    Convert Excel (.xlsx) file to Markdown format.
    
    Args:
        file_path: Path to the Excel file
        output_dir: Directory to save the generated Markdown files
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    os.makedirs(output_dir, exist_ok=True)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        md_lines = [f"# Sheet: {sheet_name}", ""]

        # Extract table data
        # First add header row
        headers = []
        header_separator = []
        if sheet.max_row > 0:
            first_row = next(sheet.iter_rows(values_only=False))
            for cell in first_row:
                value = cell.value
                headers.append(str(value) if value is not None else "")
                header_separator.append("---")
            
            md_lines.append("| " + " | ".join(headers) + " |")
            md_lines.append("| " + " | ".join(header_separator) + " |")
        
        # Now add data rows
        for row_idx, row in enumerate(sheet.iter_rows(values_only=False)):
            # Skip header row as we've already processed it
            if row_idx == 0:
                continue
                
            row_data = []
            for cell in row:
                value = cell.value
                formula = cell.value if cell.data_type == 'f' else None
                if formula:
                    row_data.append(f"`={formula}`")
                else:
                    row_data.append(str(value) if value is not None else "")
            md_lines.append("| " + " | ".join(row_data) + " |")

        # Save to markdown file
        md_file = Path(output_dir) / f"{sheet_name}.md"
        with open(md_file, "w", encoding="utf-8") as f:
            f.write("\n".join(md_lines))

    print(f"Markdown files saved to: {output_dir}")


def main():
    """Main entry point for the script."""
    parser = argparse.ArgumentParser(description='Convert Excel files to Markdown format.')
    parser.add_argument('file_path', help='Path to the Excel file (.xlsx)')
    parser.add_argument('-o', '--output-dir', default='markdown_output',
                        help='Directory to save the generated Markdown files (default: markdown_output)')
    
    args = parser.parse_args()
    
    xlsx_to_markdown(args.file_path, args.output_dir)


if __name__ == "__main__":
    main()
